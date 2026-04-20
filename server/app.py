"""
设备信息采集器 - 服务端
Flask + SQLite 实现，提供 RESTful API 和 Web 管理界面
支持：项目管理、用户管理、单位管理、设备采集、数据导出、API Key 认证
"""
import os
import io
import sys
import csv
import sqlite3
import hashlib
import secrets
import datetime
import time
import shutil
import tempfile
import zipfile
import base64
import subprocess as sp
from functools import wraps
from flask import Flask, request, jsonify, send_from_directory, send_file, g, session, redirect, url_for, render_template

from datetime import timedelta

# AES 加密（用于客户端密码密文）
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad

# AES-128 密钥（16 字节，与客户端一致）
_AES_KEY = b'DC2026SK16BYTKEY'


def aes_encrypt(plaintext):
    """AES-CBC 加密，返回 base64(IV+密文)"""
    iv = secrets.token_bytes(16)
    cipher = AES.new(_AES_KEY, AES.MODE_CBC, iv)
    ct = cipher.encrypt(pad(plaintext.encode('utf-8'), AES.block_size))
    return base64.b64encode(iv + ct).decode('ascii')

app = Flask(__name__, static_folder='static', template_folder='templates')
app.secret_key = 'device-collector-session-2026-secure'
app.permanent_session_lifetime = timedelta(days=7)

DB_PATH = os.path.join(os.path.dirname(__file__), 'data', 'devices.db')
SECRET_KEY = 'device-collector-2026'

# 登录失败限制配置
MAX_LOGIN_ATTEMPTS = 5          # 最大尝试次数
LOCKOUT_DURATION = 300          # 锁定秒数（5分钟）
login_attempts = {}             # {username: {'count': int, 'locked_until': float}}


def get_db():
    """获取数据库连接"""
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def hash_password(password):
    """密码哈希"""
    return hashlib.sha256((password + SECRET_KEY).encode()).hexdigest()


def generate_api_key():
    """生成 API Key（dc_ 前缀 + 32位随机hex）"""
    return 'dc_' + secrets.token_hex(32)


def api_response(data=None, message='success', code=200):
    """统一 API 响应格式"""
    resp = {
        'code': code,
        'message': message,
        'data': data,
        'timestamp': datetime.datetime.now().isoformat()
    }
    return jsonify(resp), code


def init_db():
    """初始化数据库表"""
    conn = get_db()
    cursor = conn.cursor()

    # 项目表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            code TEXT,
            description TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 用户表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            display_name TEXT,
            project_id INTEGER,
            role TEXT DEFAULT 'user',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id)
        )
    ''')

    # 单位信息表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS departments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            code TEXT,
            description TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id)
        )
    ''')

    # 设备信息采集表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS devices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department_id INTEGER NOT NULL,
            user_name TEXT NOT NULL,
            user_phone TEXT,
            user_position TEXT,
            computer_name TEXT,
            ip_address TEXT,
            mac_address TEXT,
            dhcp_enabled TEXT,
            os_info TEXT,
            cpu_info TEXT,
            ram_info TEXT,
            disk_info TEXT,
            motherboard_info TEXT,
            gpu_info TEXT,
            network_adapter TEXT,
            subnet_mask TEXT,
            gateway TEXT,
            dns_servers TEXT,
            collected_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments(id)
        )
    ''')

    # 操作日志表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            log_type TEXT NOT NULL,
            content TEXT NOT NULL,
            detail TEXT,
            operator TEXT,
            ip_address TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # API Key 表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS api_keys (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            api_key TEXT NOT NULL UNIQUE,
            description TEXT,
            permissions TEXT DEFAULT 'read',
            is_active INTEGER DEFAULT 1,
            last_used_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            expires_at TIMESTAMP
        )
    ''')

    # 插入示例数据（如果表为空）
    cursor.execute('SELECT COUNT(*) FROM projects')
    if cursor.fetchone()[0] == 0:
        # 示例项目
        cursor.executemany(
            'INSERT INTO projects (name, code, description) VALUES (?, ?, ?)',
            [
                ('总部设备管理', 'HQ', '总部办公设备资产管理'),
                ('研发中心设备管理', 'RD', '研发中心设备资产管理'),
                ('分公司设备管理', 'BRANCH', '分公司设备资产管理'),
            ]
        )

        # 示例用户 (默认密码: 123456)
        default_pwd = hash_password('123456')
        cursor.executemany(
            'INSERT INTO users (username, password_hash, display_name, project_id, role) VALUES (?, ?, ?, ?, ?)',
            [
                ('admin', default_pwd, '系统管理员', None, 'admin'),
                ('zhangsan', default_pwd, '张三', 1, 'user'),
                ('lisi', default_pwd, '李四', 2, 'user'),
                ('wangwu', default_pwd, '王五', 3, 'user'),
            ]
        )

        # 示例单位
        cursor.executemany(
            'INSERT INTO departments (project_id, name, code, description) VALUES (?, ?, ?, ?)',
            [
                (1, '信息技术部', 'IT-001', '信息技术部门'),
                (1, '财务部', 'FIN-001', '财务管理部门'),
                (1, '人力资源部', 'HR-001', '人力资源管理'),
                (1, '行政部', 'ADM-001', '行政管理部'),
                (2, '前端开发组', 'RD-FE', '前端开发团队'),
                (2, '后端开发组', 'RD-BE', '后端开发团队'),
                (2, '测试组', 'RD-QA', '质量保证团队'),
                (3, '华北分公司', 'BR-NORTH', '华北地区'),
                (3, '华南分公司', 'BR-SOUTH', '华南地区'),
                (3, '华东分公司', 'BR-EAST', '华东地区'),
            ]
        )

    # 插入默认 API Key（如果表为空）
    cursor.execute('SELECT COUNT(*) FROM api_keys')
    if cursor.fetchone()[0] == 0:
        default_key = generate_api_key()
        cursor.execute(
            'INSERT INTO api_keys (name, api_key, description, permissions) VALUES (?, ?, ?, ?)',
            ('默认API密钥', default_key, '系统自动生成的默认API密钥，拥有完整读写权限', 'read,write')
        )
        print(f'\n  🔑 默认 API Key: {default_key}')
        print(f'     请妥善保管，调用接口时在 Header 中传入: X-API-Key: {default_key}\n')

    conn.commit()
    conn.close()


# 确保 gunicorn 模式下也初始化数据库
init_db()


def add_log(log_type, content, detail=None, operator=None, ip_address=None):
    """写入操作日志"""
    try:
        conn = get_db()
        conn.execute(
            'INSERT INTO logs (log_type, content, detail, operator, ip_address) VALUES (?, ?, ?, ?, ?)',
            (log_type, content, detail, operator, ip_address)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"写入日志失败: {e}")


# ==================== API Key 认证 ====================

def validate_api_key(api_key_str):
    """验证 API Key 是否有效，返回 (is_valid, key_info)"""
    if not api_key_str:
        return False, None

    conn = get_db()
    row = conn.execute(
        'SELECT * FROM api_keys WHERE api_key = ? AND is_active = 1',
        (api_key_str,)
    ).fetchone()

    if not row:
        conn.close()
        return False, None

    key_info = dict(row)

    # 检查过期时间
    if key_info.get('expires_at'):
        try:
            expires = datetime.datetime.fromisoformat(key_info['expires_at'])
            if expires < datetime.datetime.now():
                conn.close()
                return False, None
        except (ValueError, TypeError):
            pass

    # 更新最后使用时间
    conn.execute(
        'UPDATE api_keys SET last_used_at = ? WHERE id = ?',
        (datetime.datetime.now().isoformat(), key_info['id'])
    )
    conn.commit()
    conn.close()

    return True, key_info


def require_api_key(permissions='read'):
    """API Key 认证装饰器

    用法:
        @require_api_key()          # 需要 read 权限
        @require_api_key('write')   # 需要 write 权限
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            api_key_str = request.headers.get('X-API-Key') or request.args.get('api_key')

            if not api_key_str:
                return api_response(message='缺少 API Key，请在请求头 X-API-Key 或参数 api_key 中传入', code=401)

            is_valid, key_info = validate_api_key(api_key_str)
            if not is_valid:
                return api_response(message='API Key 无效或已过期', code=401)

            # 检查权限
            key_perms = (key_info.get('permissions') or '').split(',')
            if permissions == 'write' and 'write' not in key_perms:
                return api_response(message='API Key 权限不足，需要 write 权限', code=403)

            # 将 key 信息存到 g 中供后续使用
            g.api_key_info = key_info

            return f(*args, **kwargs)
        return decorated_function
    return decorator


# ==================== 认证 API ====================

@app.route('/api/login', methods=['POST'])
def login():
    """用户登录（含失败次数限制）"""
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()

    if not username or not password:
        return jsonify({'error': '用户名和密码不能为空'}), 400

    # 检查是否被锁定
    attempt = login_attempts.get(username, {'count': 0, 'locked_until': 0})
    now = time.time()
    if attempt['locked_until'] > now:
        remaining = int(attempt['locked_until'] - now)
        locked_until_iso = datetime.datetime.fromtimestamp(attempt['locked_until']).isoformat()
        return jsonify({
            'error': f'登录已锁定，请 {remaining} 秒后重试',
            'locked': True,
            'locked_until': locked_until_iso
        }), 429

    # 如果锁定时间已过，重置计数
    if attempt['locked_until'] > 0 and attempt['locked_until'] <= now:
        login_attempts[username] = {'count': 0, 'locked_until': 0}

    conn = get_db()
    user = conn.execute(
        'SELECT u.*, p.name as project_name, p.code as project_code '
        'FROM users u LEFT JOIN projects p ON u.project_id = p.id '
        'WHERE u.username = ?',
        (username,)
    ).fetchone()
    conn.close()

    if not user or user['password_hash'] != hash_password(password):
        # 登录失败，增加计数
        attempt = login_attempts.get(username, {'count': 0, 'locked_until': 0})
        attempt['count'] += 1
        attempts_left = MAX_LOGIN_ATTEMPTS - attempt['count']

        if attempt['count'] >= MAX_LOGIN_ATTEMPTS:
            # 锁定账户
            attempt['locked_until'] = now + LOCKOUT_DURATION
            login_attempts[username] = attempt
            # 记录安全日志
            add_log('LOGIN_LOCKOUT',
                    f'登录锁定: {username}',
                    f'连续失败 {attempt["count"]} 次，锁定 {LOCKOUT_DURATION} 秒',
                    ip_address=request.remote_addr)
            return jsonify({
                'error': f'连续登录失败 {attempt["count"]} 次，账户已锁定 {LOCKOUT_DURATION // 60} 分钟',
                'locked': True,
                'locked_until': datetime.datetime.fromtimestamp(attempt['locked_until']).isoformat()
            }), 429

        login_attempts[username] = attempt

        # 记录登录失败日志
        add_log('LOGIN_FAILED',
                f'登录失败: {username}',
                f'第 {attempt["count"]} 次失败，剩余 {attempts_left} 次',
                ip_address=request.remote_addr)

        return jsonify({
            'error': '用户名或密码错误',
            'attempts_left': attempts_left
        }), 401

    # 登录成功，清除失败计数
    login_attempts.pop(username, None)

    # 设置 session
    session['user_id'] = user['id']
    session['username'] = user['username']
    session['display_name'] = user['display_name']
    session['role'] = user['role']
    session['project_id'] = user['project_id']
    session.permanent = True

    # 记录登录日志
    add_log('USER_LOGIN',
            f'用户登录: {user["display_name"] or user["username"]}',
            f'用户名:{user["username"]}, 角色:{user["role"]}',
            operator=user['username'],
            ip_address=request.remote_addr)

    return jsonify({
        'id': user['id'],
        'username': user['username'],
        'display_name': user['display_name'],
        'role': user['role'],
        'project_id': user['project_id'],
        'project_name': user['project_name'],
    })


@app.route('/api/logout', methods=['POST'])
def logout():
    """用户登出"""
    username = session.get('username', '未知')
    session.clear()
    return jsonify({'message': '已登出'})


@app.route('/api/current-user', methods=['GET'])
def current_user():
    """获取当前登录用户信息"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    return jsonify({
        'id': session.get('user_id'),
        'username': session.get('username'),
        'display_name': session.get('display_name'),
        'role': session.get('role'),
        'project_id': session.get('project_id'),
    })


# ==================== 项目 API ====================

@app.route('/api/projects', methods=['GET'])
def get_projects():
    """获取所有项目"""
    conn = get_db()
    rows = conn.execute('''
        SELECT p.*,
            (SELECT COUNT(*) FROM users WHERE project_id = p.id) as user_count,
            (SELECT COUNT(*) FROM departments WHERE project_id = p.id) as dept_count,
            (SELECT COUNT(*) FROM devices d JOIN departments dept ON d.department_id = dept.id WHERE dept.project_id = p.id) as device_count
        FROM projects p ORDER BY p.id
    ''').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/projects', methods=['POST'])
def add_project():
    """添加项目"""
    data = request.json
    name = data.get('name', '').strip()
    code = data.get('code', '').strip()
    description = data.get('description', '').strip()

    if not name:
        return jsonify({'error': '项目名称不能为空'}), 400

    conn = get_db()
    try:
        cursor = conn.execute(
            'INSERT INTO projects (name, code, description) VALUES (?, ?, ?)',
            (name, code, description)
        )
        conn.commit()
        return jsonify({'message': '添加成功', 'id': cursor.lastrowid}), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': '项目名称已存在'}), 409
    finally:
        conn.close()


@app.route('/api/projects/<int:project_id>', methods=['PUT'])
def update_project(project_id):
    """更新项目"""
    data = request.json
    name = data.get('name', '').strip()
    code = data.get('code', '').strip()
    description = data.get('description', '').strip()

    if not name:
        return jsonify({'error': '项目名称不能为空'}), 400

    conn = get_db()
    try:
        conn.execute(
            'UPDATE projects SET name=?, code=?, description=? WHERE id=?',
            (name, code, description, project_id)
        )
        conn.commit()
        return jsonify({'message': '更新成功'})
    except sqlite3.IntegrityError:
        return jsonify({'error': '项目名称已存在'}), 409
    finally:
        conn.close()


@app.route('/api/projects/<int:project_id>', methods=['DELETE'])
def delete_project(project_id):
    """删除项目"""
    conn = get_db()
    # 检查是否有关联数据
    user_count = conn.execute('SELECT COUNT(*) FROM users WHERE project_id=?', (project_id,)).fetchone()[0]
    dept_count = conn.execute('SELECT COUNT(*) FROM departments WHERE project_id=?', (project_id,)).fetchone()[0]
    if user_count > 0 or dept_count > 0:
        return jsonify({'error': f'该项目下还有 {user_count} 个用户和 {dept_count} 个单位，请先删除关联数据'}), 400
    conn.execute('DELETE FROM projects WHERE id = ?', (project_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': '删除成功'})


# ==================== 用户 API ====================

@app.route('/api/users', methods=['GET'])
def get_users():
    """获取用户列表，支持按项目筛选"""
    project_id = request.args.get('project_id')
    conn = get_db()
    query = '''
        SELECT u.*, p.name as project_name, p.code as project_code
        FROM users u
        LEFT JOIN projects p ON u.project_id = p.id
        WHERE 1=1
    '''
    params = []
    if project_id:
        query += ' AND u.project_id = ?'
        params.append(project_id)
    query += ' ORDER BY u.id'
    rows = conn.execute(query, params).fetchall()
    conn.close()
    # 不返回密码
    result = []
    for r in rows:
        d = dict(r)
        d.pop('password_hash', None)
        result.append(d)
    return jsonify(result)


@app.route('/api/users', methods=['POST'])
def add_user():
    """添加用户"""
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    display_name = data.get('display_name', '').strip()
    project_id = data.get('project_id')
    role = data.get('role', 'user').strip()

    if not username or not password:
        return jsonify({'error': '用户名和密码不能为空'}), 400

    if role not in ('admin', 'user'):
        role = 'user'

    conn = get_db()
    try:
        cursor = conn.execute(
            'INSERT INTO users (username, password_hash, display_name, project_id, role) VALUES (?, ?, ?, ?, ?)',
            (username, hash_password(password), display_name, project_id if project_id else None, role)
        )
        conn.commit()
        return jsonify({'message': '添加成功', 'id': cursor.lastrowid}), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': '用户名已存在'}), 409
    finally:
        conn.close()


@app.route('/api/users/<int:user_id>', methods=['PUT'])
def update_user(user_id):
    """更新用户"""
    data = request.json
    display_name = data.get('display_name', '').strip()
    project_id = data.get('project_id')
    role = data.get('role', 'user').strip()
    password = data.get('password', '').strip()

    if role not in ('admin', 'user'):
        role = 'user'

    conn = get_db()
    if password:
        conn.execute(
            'UPDATE users SET display_name=?, project_id=?, role=?, password_hash=? WHERE id=?',
            (display_name, project_id if project_id else None, role, hash_password(password), user_id)
        )
    else:
        conn.execute(
            'UPDATE users SET display_name=?, project_id=?, role=? WHERE id=?',
            (display_name, project_id if project_id else None, role, user_id)
        )
    conn.commit()
    conn.close()
    return jsonify({'message': '更新成功'})


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    """删除用户"""
    conn = get_db()
    conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': '删除成功'})


@app.route('/api/users/reset-password/<int:user_id>', methods=['POST'])
def reset_user_password(user_id):
    """重置用户密码"""
    data = request.json
    new_password = data.get('password', '123456').strip()
    conn = get_db()
    conn.execute('UPDATE users SET password_hash=? WHERE id=?', (hash_password(new_password), user_id))
    conn.commit()
    conn.close()
    return jsonify({'message': f'密码已重置为: {new_password}'})


# ==================== 单位 API ====================

@app.route('/api/departments', methods=['GET'])
def get_departments():
    """获取单位列表，支持按项目筛选"""
    project_id = request.args.get('project_id')
    conn = get_db()
    query = 'SELECT d.*, p.name as project_name FROM departments d LEFT JOIN projects p ON d.project_id = p.id WHERE 1=1'
    params = []
    if project_id:
        query += ' AND d.project_id = ?'
        params.append(project_id)
    query += ' ORDER BY d.id'
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/departments', methods=['POST'])
def add_department():
    """添加单位"""
    data = request.json
    name = data.get('name', '').strip()
    code = data.get('code', '').strip()
    description = data.get('description', '').strip()
    project_id = data.get('project_id')

    if not name:
        return jsonify({'error': '单位名称不能为空'}), 400
    if not project_id:
        return jsonify({'error': '请选择所属项目'}), 400

    conn = get_db()
    try:
        conn.execute(
            'INSERT INTO departments (project_id, name, code, description) VALUES (?, ?, ?, ?)',
            (project_id, name, code, description)
        )
        conn.commit()
        return jsonify({'message': '添加成功'}), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': '单位名称已存在'}), 409
    finally:
        conn.close()


@app.route('/api/departments/batch', methods=['POST'])
def batch_import_departments():
    """批量导入单位"""
    project_id = request.form.get('project_id') or (request.json and request.json.get('project_id'))
    if not project_id:
        return jsonify({'error': '请指定所属项目'}), 400
    project_id = int(project_id)

    if 'file' in request.files:
        file = request.files['file']
        if not file.filename:
            return jsonify({'error': '未选择文件'}), 400

        filename = file.filename.lower()
        departments = []

        if filename.endswith('.csv'):
            stream = io.StringIO(file.read().decode('utf-8-sig'))
            reader = csv.reader(stream)
            for row in reader:
                if not row or not row[0].strip():
                    continue
                if row[0].strip() in ('单位名称', '名称', 'name', 'Name', '部门'):
                    continue
                name = row[0].strip()
                code = row[1].strip() if len(row) > 1 else ''
                description = row[2].strip() if len(row) > 2 else ''
                departments.append((project_id, name, code, description))

        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            try:
                import openpyxl
                file.stream.seek(0)
                wb = openpyxl.load_workbook(io.BytesIO(file.read()))
                ws = wb.active
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if not row or not row[0] or not str(row[0]).strip():
                        continue
                    if i == 0 and str(row[0]).strip() in ('单位名称', '名称', 'name', 'Name', '部门'):
                        continue
                    name = str(row[0]).strip()
                    code = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    description = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                    departments.append((project_id, name, code, description))
            except Exception as e:
                return jsonify({'error': f'Excel 解析失败: {str(e)}'}), 400
        else:
            return jsonify({'error': '不支持的文件格式，请使用 CSV 或 Excel 文件'}), 400
    else:
        data = request.json
        if not data or not isinstance(data, list):
            return jsonify({'error': '请提供 JSON 数组或上传 CSV/Excel 文件'}), 400
        departments = []
        for item in data:
            name = item.get('name', '').strip() if isinstance(item, dict) else str(item).strip()
            code = item.get('code', '').strip() if isinstance(item, dict) else ''
            description = item.get('description', '').strip() if isinstance(item, dict) else ''
            if name:
                departments.append((project_id, name, code, description))

    if not departments:
        return jsonify({'error': '未找到有效的单位数据'}), 400

    conn = get_db()
    success_count = 0
    skip_count = 0
    errors = []

    for pid, name, code, description in departments:
        try:
            conn.execute('INSERT INTO departments (project_id, name, code, description) VALUES (?, ?, ?, ?)',
                         (pid, name, code, description))
            success_count += 1
        except sqlite3.IntegrityError:
            skip_count += 1
            errors.append(f'"{name}" 已存在，已跳过')

    conn.commit()
    conn.close()

    result = {
        'message': f'导入完成：成功 {success_count} 个，跳过 {skip_count} 个',
        'success_count': success_count,
        'skip_count': skip_count,
    }
    if errors:
        result['details'] = errors
    return jsonify(result)


@app.route('/api/departments/template', methods=['GET'])
def download_template():
    """下载单位导入模板"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['单位名称', '单位编码', '描述'])
    writer.writerow(['示例部门A', 'DEPT-001', '这是一个示例部门'])
    writer.writerow(['示例部门B', 'DEPT-002', '另一个示例部门'])

    mem = io.BytesIO()
    mem.write(output.getvalue().encode('utf-8-sig'))
    mem.seek(0)

    return send_file(mem, mimetype='text/csv', as_attachment=True, download_name='单位导入模板.csv')


@app.route('/api/departments/<int:dept_id>', methods=['PUT'])
def update_department(dept_id):
    """编辑单位"""
    data = request.json
    name = data.get('name', '').strip()
    code = data.get('code', '').strip()
    description = data.get('description', '').strip()
    project_id = data.get('project_id')

    if not name:
        return jsonify({'error': '单位名称不能为空'}), 400

    conn = get_db()
    try:
        conn.execute(
            'UPDATE departments SET name=?, code=?, description=?, project_id=? WHERE id=?',
            (name, code, description, project_id if project_id else None, dept_id)
        )
        conn.commit()
        add_log('DEPT_EDIT',
                f'单位编辑 (ID:{dept_id})',
                f'名称:{name}, 编码:{code}, 项目ID:{project_id}',
                ip_address=request.remote_addr)
        return jsonify({'message': '更新成功'})
    except sqlite3.IntegrityError:
        return jsonify({'error': '单位名称已存在'}), 409
    finally:
        conn.close()


@app.route('/api/departments/<int:dept_id>', methods=['DELETE'])
def delete_department(dept_id):
    """删除单位"""
    conn = get_db()
    conn.execute('DELETE FROM departments WHERE id = ?', (dept_id,))
    conn.commit()
    conn.close()
    add_log('DEPT_DELETE',
            f'单位删除 (ID:{dept_id})',
            ip_address=request.remote_addr)
    return jsonify({'message': '删除成功'})


# ==================== 设备 API ====================

@app.route('/api/devices/check-duplicates', methods=['GET'])
def check_device_duplicates():
    """检查设备IP/MAC重复"""
    conn = get_db()
    
    # 查找IP重复的设备
    ip_dups = conn.execute('''
        SELECT d1.id, d1.computer_name, d1.user_name, d1.ip_address, d1.mac_address,
               dept.name as department_name, p.name as project_name,
               COUNT(*) as dup_count
        FROM devices d1
        LEFT JOIN departments dept ON d1.department_id = dept.id
        LEFT JOIN projects p ON dept.project_id = p.id
        INNER JOIN devices d2 ON d1.ip_address = d2.ip_address AND d1.id != d2.id
        WHERE d1.ip_address IS NOT NULL AND d1.ip_address != '' 
              AND d1.ip_address != '未知' AND d1.ip_address != '0.0.0.0'
        GROUP BY d1.ip_address, d1.id
        ORDER BY d1.ip_address
    ''').fetchall()
    
    # 查找MAC重复的设备
    mac_dups = conn.execute('''
        SELECT d1.id, d1.computer_name, d1.user_name, d1.ip_address, d1.mac_address,
               dept.name as department_name, p.name as project_name,
               COUNT(*) as dup_count
        FROM devices d1
        LEFT JOIN departments dept ON d1.department_id = dept.id
        LEFT JOIN projects p ON dept.project_id = p.id
        INNER JOIN devices d2 ON d1.mac_address = d2.mac_address AND d1.id != d2.id
        WHERE d1.mac_address IS NOT NULL AND d1.mac_address != '' 
              AND d1.mac_address != '未知' AND d1.mac_address != '00:00:00:00:00:00'
        GROUP BY d1.mac_address, d1.id
        ORDER BY d1.mac_address
    ''').fetchall()
    
    conn.close()
    
    # 按IP分组
    ip_groups = {}
    for r in ip_dups:
        ip = r['ip_address']
        if ip not in ip_groups:
            ip_groups[ip] = []
        ip_groups[ip].append(dict(r))
    
    # 按MAC分组
    mac_groups = {}
    for r in mac_dups:
        mac = r['mac_address']
        if mac not in mac_groups:
            mac_groups[mac] = []
        mac_groups[mac].append(dict(r))
    
    return jsonify({
        'ip_duplicates': {k: v for k, v in ip_groups.items()},
        'mac_duplicates': {k: v for k, v in mac_groups.items()},
        'ip_duplicate_count': len(ip_groups),
        'mac_duplicate_count': len(mac_groups),
        'total_duplicate_devices': len(set(
            [r['id'] for r in ip_dups] + [r['id'] for r in mac_dups]
        ))
    })


@app.route('/api/devices', methods=['POST'])
def submit_device():
    """客户端提交设备信息"""
    data = request.json
    required = ['department_id', 'user_name']
    for field in required:
        if not data.get(field):
            return jsonify({'error': f'缺少必填字段: {field}'}), 400

    ip_address = data.get('ip_address', '')
    mac_address = data.get('mac_address', '')
    force = data.get('force', False)
    username = data.get('_username', '未知')
    computer_name = data.get('computer_name', '未知')

    conn = get_db()

    # 检查 IP/MAC 重复
    duplicates = []
    if ip_address and ip_address not in ('未知', '0.0.0.0'):
        ip_dup = conn.execute('''
            SELECT d.id, d.computer_name, d.user_name, d.ip_address, d.mac_address,
                   dept.name as department_name
            FROM devices d
            LEFT JOIN departments dept ON d.department_id = dept.id
            WHERE d.ip_address = ? AND d.ip_address != '未知'
        ''', (ip_address,)).fetchall()
        for r in ip_dup:
            duplicates.append({
                'type': 'IP地址重复',
                'field': 'ip_address',
                'value': ip_address,
                'device_id': r['id'],
                'computer_name': r['computer_name'],
                'user_name': r['user_name'],
                'department_name': r['department_name'],
                'mac_address': r['mac_address'],
            })

    if mac_address and mac_address not in ('未知', '00:00:00:00:00:00'):
        mac_dup = conn.execute('''
            SELECT d.id, d.computer_name, d.user_name, d.ip_address, d.mac_address,
                   dept.name as department_name
            FROM devices d
            LEFT JOIN departments dept ON d.department_id = dept.id
            WHERE d.mac_address = ? AND d.mac_address != '未知'
        ''', (mac_address,)).fetchall()
        for r in mac_dup:
            duplicates.append({
                'type': 'MAC地址重复',
                'field': 'mac_address',
                'value': mac_address,
                'device_id': r['id'],
                'computer_name': r['computer_name'],
                'user_name': r['user_name'],
                'department_name': r['department_name'],
                'ip_address': r['ip_address'],
            })

    if duplicates and not force:
        # 有重复但用户未确认，返回重复信息让客户端确认
        conn.close()
        add_log('DUPLICATE_WARNING',
                f'设备提交检测到IP/MAC重复（用户未确认）',
                f'电脑:{computer_name}, IP:{ip_address}, MAC:{mac_address}, 操作人:{username}',
                operator=username)
        return jsonify({
            'duplicate': True,
            'message': '检测到IP或MAC地址与已有设备重复',
            'duplicates': duplicates
        }), 409

    if duplicates and force:
        # 用户确认强制提交，记录日志
        dup_info = '; '.join(
            f"{d['type']}:{d['value']}(已存在设备ID:{d['device_id']}, 电脑:{d['computer_name']}, 使用人:{d['user_name']})"
            for d in duplicates
        )
        add_log('DUPLICATE_CONFIRMED',
                f'设备提交（用户确认覆盖重复）',
                f'新电脑:{computer_name}, IP:{ip_address}, MAC:{mac_address}, 重复详情:[{dup_info}], 操作人:{username}',
                operator=username)

    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO devices (
            department_id, user_name, user_phone, user_position,
            computer_name, ip_address, mac_address, dhcp_enabled,
            os_info, cpu_info, ram_info, disk_info,
            motherboard_info, gpu_info, network_adapter,
            subnet_mask, gateway, dns_servers
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data.get('department_id'),
        data.get('user_name'),
        data.get('user_phone'),
        data.get('user_position'),
        data.get('computer_name'),
        data.get('ip_address'),
        data.get('mac_address'),
        data.get('dhcp_enabled'),
        data.get('os_info'),
        data.get('cpu_info'),
        data.get('ram_info'),
        data.get('disk_info'),
        data.get('motherboard_info'),
        data.get('gpu_info'),
        data.get('network_adapter'),
        data.get('subnet_mask'),
        data.get('gateway'),
        data.get('dns_servers'),
    ))
    conn.commit()
    device_id = cursor.lastrowid
    conn.close()

    # 记录提交日志
    add_log('DEVICE_SUBMIT',
            f'设备信息提交成功',
            f'设备ID:{device_id}, 电脑:{computer_name}, IP:{ip_address}, MAC:{mac_address}, 使用人:{data.get("user_name")}, 操作人:{username}',
            operator=username)

    return jsonify({'message': '提交成功', 'id': device_id}), 201


@app.route('/api/devices', methods=['GET'])
def get_devices():
    """获取设备列表，支持按项目/单位筛选"""
    project_id = request.args.get('project_id')
    dept_id = request.args.get('department_id')
    keyword = request.args.get('keyword', '').strip()

    conn = get_db()
    query = '''
        SELECT d.*, dept.name as department_name, dept.code as department_code,
               dept.project_id, p.name as project_name
        FROM devices d
        LEFT JOIN departments dept ON d.department_id = dept.id
        LEFT JOIN projects p ON dept.project_id = p.id
        WHERE 1=1
    '''
    params = []

    if project_id:
        query += ' AND dept.project_id = ?'
        params.append(project_id)

    if dept_id:
        query += ' AND d.department_id = ?'
        params.append(dept_id)

    if keyword:
        query += ' AND (d.user_name LIKE ? OR d.computer_name LIKE ? OR d.ip_address LIKE ?)'
        kw = f'%{keyword}%'
        params.extend([kw, kw, kw])

    query += ' ORDER BY d.collected_at DESC'

    rows = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/devices/<int:device_id>', methods=['PUT'])
def update_device(device_id):
    """编辑设备记录"""
    data = request.json
    conn = get_db()

    # 构建可编辑字段
    editable_fields = [
        'department_id', 'user_name', 'user_phone', 'user_position',
        'computer_name', 'ip_address', 'mac_address', 'dhcp_enabled',
        'os_info', 'cpu_info', 'ram_info', 'disk_info',
        'motherboard_info', 'gpu_info', 'network_adapter',
        'subnet_mask', 'gateway', 'dns_servers'
    ]

    updates = []
    values = []
    for field in editable_fields:
        if field in data:
            updates.append(f'{field} = ?')
            values.append(data[field])

    if not updates:
        return jsonify({'error': '没有需要更新的字段'}), 400

    values.append(device_id)
    sql = f"UPDATE devices SET {', '.join(updates)} WHERE id = ?"
    conn.execute(sql, values)
    conn.commit()

    # 记录编辑日志
    changed = ', '.join(f"{f}={data.get(f, '')}" for f in editable_fields if f in data)
    add_log('DEVICE_EDIT',
            f'设备记录编辑 (ID:{device_id})',
            f'修改字段: {changed}',
            ip_address=request.remote_addr)

    conn.close()
    return jsonify({'message': '更新成功'})


@app.route('/api/devices/<int:device_id>', methods=['GET'])
def get_device(device_id):
    """获取单个设备详情"""
    conn = get_db()
    row = conn.execute('''
        SELECT d.*, dept.name as department_name, dept.code as department_code,
               dept.project_id, p.name as project_name
        FROM devices d
        LEFT JOIN departments dept ON d.department_id = dept.id
        LEFT JOIN projects p ON dept.project_id = p.id
        WHERE d.id = ?
    ''', (device_id,)).fetchone()
    conn.close()
    if row:
        return jsonify(dict(row))
    return jsonify({'error': '设备不存在'}), 404


@app.route('/api/devices/<int:device_id>', methods=['DELETE'])
def delete_device(device_id):
    """删除设备记录"""
    conn = get_db()
    # 先查询设备信息用于日志
    row = conn.execute(
        'SELECT d.*, dept.name as department_name FROM devices d LEFT JOIN departments dept ON d.department_id = dept.id WHERE d.id = ?',
        (device_id,)
    ).fetchone()
    device_info = dict(row) if row else {}
    conn.execute('DELETE FROM devices WHERE id = ?', (device_id,))
    conn.commit()
    conn.close()

    # 记录删除日志
    add_log('DEVICE_DELETE',
            f'设备记录删除 (ID:{device_id})',
            f'电脑:{device_info.get("computer_name","?")}, IP:{device_info.get("ip_address","?")}, MAC:{device_info.get("mac_address","?")}, 使用人:{device_info.get("user_name","?")}, 单位:{device_info.get("department_name","?")}',
            ip_address=request.remote_addr)

    return jsonify({'message': '删除成功'})


@app.route('/api/devices/export', methods=['GET'])
def export_devices():
    """导出设备列表"""
    export_format = request.args.get('format', 'xlsx').lower()
    project_id = request.args.get('project_id')
    dept_id = request.args.get('department_id')
    keyword = request.args.get('keyword', '').strip()

    conn = get_db()
    query = '''
        SELECT d.*, dept.name as department_name, dept.code as department_code,
               dept.project_id, p.name as project_name
        FROM devices d
        LEFT JOIN departments dept ON d.department_id = dept.id
        LEFT JOIN projects p ON dept.project_id = p.id
        WHERE 1=1
    '''
    params = []

    if project_id:
        query += ' AND dept.project_id = ?'
        params.append(project_id)
    if dept_id:
        query += ' AND d.department_id = ?'
        params.append(dept_id)
    if keyword:
        query += ' AND (d.user_name LIKE ? OR d.computer_name LIKE ? OR d.ip_address LIKE ?)'
        kw = f'%{keyword}%'
        params.extend([kw, kw, kw])

    query += ' ORDER BY d.collected_at DESC'
    rows = conn.execute(query, params).fetchall()
    conn.close()

    columns = [
        ('ID', 'id'),
        ('所属项目', 'project_name'),
        ('所属单位', 'department_name'),
        ('单位编码', 'department_code'),
        ('使用人', 'user_name'),
        ('联系电话', 'user_phone'),
        ('安装位置', 'user_position'),
        ('电脑名称', 'computer_name'),
        ('IP 地址', 'ip_address'),
        ('MAC 地址', 'mac_address'),
        ('自动获取IP(DHCP)', 'dhcp_enabled'),
        ('子网掩码', 'subnet_mask'),
        ('默认网关', 'gateway'),
        ('DNS 服务器', 'dns_servers'),
        ('网卡', 'network_adapter'),
        ('操作系统', 'os_info'),
        ('CPU', 'cpu_info'),
        ('内存', 'ram_info'),
        ('硬盘', 'disk_info'),
        ('主板', 'motherboard_info'),
        ('显卡', 'gpu_info'),
        ('采集时间', 'collected_at'),
    ]

    if export_format == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([col[0] for col in columns])
        for row in rows:
            r = dict(row)
            writer.writerow([r.get(col[1], '') or '' for col in columns])

        mem = io.BytesIO()
        mem.write(output.getvalue().encode('utf-8-sig'))
        mem.seek(0)
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(mem, mimetype='text/csv', as_attachment=True,
                         download_name=f'设备列表_{timestamp}.csv')

    else:  # xlsx
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '设备列表'

        header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='1A73E8', end_color='1A73E8', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0'),
        )

        for col_idx, (header, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        data_font = Font(name='微软雅黑', size=10)
        data_align = Alignment(vertical='center', wrap_text=True)
        for row_idx, row in enumerate(rows, 2):
            r = dict(row)
            for col_idx, (_, key) in enumerate(columns, 1):
                val = r.get(key, '') or ''
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

        col_widths = {
            'ID': 6, '所属项目': 16, '所属单位': 14, '单位编码': 10,
            '使用人': 10, '联系电话': 14, '安装位置': 14, '电脑名称': 18,
            'IP 地址': 14, 'MAC 地址': 18, '自动获取IP(DHCP)': 14,
            '子网掩码': 14, '默认网关': 14, 'DNS 服务器': 16,
            '网卡': 30, '操作系统': 35, 'CPU': 35, '内存': 10,
            '硬盘': 35, '主板': 30, '显卡': 20, '采集时间': 18,
        }
        for col_idx, (header, _) in enumerate(columns, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = \
                col_widths.get(header, 15)

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        mem = io.BytesIO()
        wb.save(mem)
        mem.seek(0)
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(
            mem,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'设备列表_{timestamp}.xlsx'
        )


# ==================== 日志 API ====================

@app.route('/api/logs', methods=['GET'])
def get_logs():
    """获取日志列表，支持按类型筛选和关键词搜索"""
    log_type = request.args.get('type', '').strip()
    keyword = request.args.get('keyword', '').strip()
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))

    conn = get_db()
    query = 'SELECT * FROM logs WHERE 1=1'
    params = []

    if log_type:
        query += ' AND log_type = ?'
        params.append(log_type)

    if keyword:
        query += ' AND (content LIKE ? OR detail LIKE ? OR operator LIKE ?)'
        kw = f'%{keyword}%'
        params.extend([kw, kw, kw])

    # 总数
    count_query = query.replace('SELECT *', 'SELECT COUNT(*)')
    total = conn.execute(count_query, params).fetchone()[0]

    query += ' ORDER BY created_at DESC'
    query += f' LIMIT {per_page} OFFSET {(page - 1) * per_page}'

    rows = conn.execute(query, params).fetchall()
    conn.close()

    return jsonify({
        'total': total,
        'page': page,
        'per_page': per_page,
        'data': [dict(r) for r in rows]
    })


@app.route('/api/logs/types', methods=['GET'])
def get_log_types():
    """获取所有日志类型"""
    conn = get_db()
    rows = conn.execute('SELECT DISTINCT log_type FROM logs ORDER BY log_type').fetchall()
    conn.close()
    return jsonify([r['log_type'] for r in rows])


@app.route('/api/logs/<int:log_id>', methods=['DELETE'])
def delete_log(log_id):
    """删除日志"""
    conn = get_db()
    conn.execute('DELETE FROM logs WHERE id = ?', (log_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': '删除成功'})


@app.route('/api/logs/clear', methods=['POST'])
def clear_logs():
    """清空日志"""
    days = request.json.get('days', 0) if request.json else 0
    conn = get_db()
    if days > 0:
        conn.execute("DELETE FROM logs WHERE created_at < datetime('now', ?)", (f'-{days} days',))
    else:
        conn.execute('DELETE FROM logs')
    conn.commit()
    deleted = conn.total_changes
    conn.close()
    return jsonify({'message': f'已清理日志'})


# ==================== 统计 API ====================

@app.route('/api/stats', methods=['GET'])
def get_stats():
    """获取统计信息"""
    conn = get_db()
    project_count = conn.execute('SELECT COUNT(*) FROM projects').fetchone()[0]
    user_count = conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]
    dept_count = conn.execute('SELECT COUNT(*) FROM departments').fetchone()[0]
    device_count = conn.execute('SELECT COUNT(*) FROM devices').fetchone()[0]
    today_count = conn.execute(
        "SELECT COUNT(*) FROM devices WHERE DATE(collected_at) = DATE('now')"
    ).fetchone()[0]

    # 各项目设备数量
    project_stats = conn.execute('''
        SELECT p.name, COALESCE(COUNT(d.id), 0) as count
        FROM projects p
        LEFT JOIN departments dept ON dept.project_id = p.id
        LEFT JOIN devices d ON d.department_id = dept.id
        GROUP BY p.id
        ORDER BY count DESC
    ''').fetchall()

    conn.close()
    return jsonify({
        'project_count': project_count,
        'user_count': user_count,
        'department_count': dept_count,
        'device_count': device_count,
        'today_count': today_count,
        'project_stats': [dict(r) for r in project_stats]
    })


# ==================== API Key 管理 ====================

@app.route('/api/api-keys', methods=['GET'])
def get_api_keys():
    """获取所有 API Key 列表"""
    conn = get_db()
    rows = conn.execute('''
        SELECT id, name, api_key, description, permissions, is_active, last_used_at, created_at, expires_at
        FROM api_keys ORDER BY created_at DESC
    ''').fetchall()
    conn.close()
    # 脱敏：只显示前8位和后4位
    result = []
    for r in rows:
        d = dict(r)
        key = d.get('api_key', '')
        if len(key) > 12:
            d['api_key_masked'] = key[:8] + '...' + key[-4:]
        else:
            d['api_key_masked'] = key
        d.pop('api_key', None)  # 不返回完整 key
        result.append(d)
    return jsonify(result)


@app.route('/api/api-keys', methods=['POST'])
def create_api_key():
    """创建新的 API Key"""
    data = request.json
    name = data.get('name', '').strip()
    description = data.get('description', '').strip()
    permissions = data.get('permissions', 'read').strip()
    expires_at = data.get('expires_at', '').strip() or None

    if not name:
        return jsonify({'error': 'API Key 名称不能为空'}), 400

    if permissions not in ('read', 'write', 'read,write', 'read,write,admin'):
        permissions = 'read'

    new_key = generate_api_key()
    conn = get_db()
    try:
        cursor = conn.execute(
            'INSERT INTO api_keys (name, api_key, description, permissions, expires_at) VALUES (?, ?, ?, ?, ?)',
            (name, new_key, description, permissions, expires_at)
        )
        conn.commit()
        add_log('API_KEY_CREATE',
                f'创建 API Key: {name}',
                f'权限: {permissions}',
                ip_address=request.remote_addr)
        return jsonify({'message': '创建成功', 'id': cursor.lastrowid, 'api_key': new_key}), 201
    except Exception as e:
        return jsonify({'error': f'创建失败: {str(e)}'}), 400
    finally:
        conn.close()


@app.route('/api/api-keys/<int:key_id>', methods=['PUT'])
def update_api_key(key_id):
    """更新 API Key"""
    data = request.json
    conn = get_db()
    updates = []
    values = []

    if 'name' in data:
        updates.append('name = ?')
        values.append(data['name'].strip())
    if 'description' in data:
        updates.append('description = ?')
        values.append(data['description'].strip())
    if 'permissions' in data:
        updates.append('permissions = ?')
        values.append(data['permissions'].strip())
    if 'is_active' in data:
        updates.append('is_active = ?')
        values.append(1 if data['is_active'] else 0)
    if 'expires_at' in data:
        updates.append('expires_at = ?')
        values.append(data['expires_at'] or None)

    if not updates:
        return jsonify({'error': '没有需要更新的字段'}), 400

    values.append(key_id)
    conn.execute(f"UPDATE api_keys SET {', '.join(updates)} WHERE id = ?", values)
    conn.commit()
    conn.close()

    add_log('API_KEY_UPDATE',
            f'更新 API Key (ID:{key_id})',
            ip_address=request.remote_addr)
    return jsonify({'message': '更新成功'})


@app.route('/api/api-keys/<int:key_id>', methods=['DELETE'])
def delete_api_key(key_id):
    """删除 API Key"""
    conn = get_db()
    conn.execute('DELETE FROM api_keys WHERE id = ?', (key_id,))
    conn.commit()
    conn.close()

    add_log('API_KEY_DELETE',
            f'删除 API Key (ID:{key_id})',
            ip_address=request.remote_addr)
    return jsonify({'message': '删除成功'})


@app.route('/api/api-keys/<int:key_id>/regenerate', methods=['POST'])
def regenerate_api_key(key_id):
    """重新生成 API Key"""
    new_key = generate_api_key()
    conn = get_db()
    conn.execute('UPDATE api_keys SET api_key = ? WHERE id = ?', (new_key, key_id))
    conn.commit()
    conn.close()

    add_log('API_KEY_REGENERATE',
            f'重新生成 API Key (ID:{key_id})',
            ip_address=request.remote_addr)
    return jsonify({'message': '重新生成成功', 'api_key': new_key})


# ==================== V1 标准化 API 接口 ====================
# 供外部业务系统调用，需要 API Key 认证

@app.route('/api/v1/devices', methods=['GET'])
@require_api_key('read')
def v1_get_devices():
    """获取设备列表（标准化接口）

    查询参数:
        project_id    - 按项目ID筛选
        department_id - 按单位ID筛选
        keyword       - 关键词搜索（姓名/电脑名/IP）
        page          - 页码，默认1
        per_page      - 每页数量，默认20
    """
    project_id = request.args.get('project_id')
    dept_id = request.args.get('department_id')
    keyword = request.args.get('keyword', '').strip()
    page = max(1, int(request.args.get('page', 1)))
    per_page = min(100, max(1, int(request.args.get('per_page', 20))))

    conn = get_db()
    query = '''
        SELECT d.*, dept.name as department_name, dept.code as department_code,
               dept.project_id, p.name as project_name
        FROM devices d
        LEFT JOIN departments dept ON d.department_id = dept.id
        LEFT JOIN projects p ON dept.project_id = p.id
        WHERE 1=1
    '''
    count_query = '''
        SELECT COUNT(*) as total
        FROM devices d
        LEFT JOIN departments dept ON d.department_id = dept.id
        LEFT JOIN projects p ON dept.project_id = p.id
        WHERE 1=1
    '''
    params = []

    if project_id:
        query += ' AND dept.project_id = ?'
        params.append(project_id)
    if dept_id:
        query += ' AND d.department_id = ?'
        params.append(dept_id)
    if keyword:
        query += ' AND (d.user_name LIKE ? OR d.computer_name LIKE ? OR d.ip_address LIKE ?)'
        kw = f'%{keyword}%'
        params.extend([kw, kw, kw])

    total = conn.execute(count_query, params).fetchone()[0]
    query += ' ORDER BY d.collected_at DESC'
    query += f' LIMIT {per_page} OFFSET {(page - 1) * per_page}'

    rows = conn.execute(query, params).fetchall()
    conn.close()

    return api_response(data={
        'items': [dict(r) for r in rows],
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page
    })


@app.route('/api/v1/devices/<int:device_id>', methods=['GET'])
@require_api_key('read')
def v1_get_device(device_id):
    """获取单个设备详情"""
    conn = get_db()
    row = conn.execute('''
        SELECT d.*, dept.name as department_name, dept.code as department_code,
               dept.project_id, p.name as project_name
        FROM devices d
        LEFT JOIN departments dept ON d.department_id = dept.id
        LEFT JOIN projects p ON dept.project_id = p.id
        WHERE d.id = ?
    ''', (device_id,)).fetchone()
    conn.close()

    if not row:
        return api_response(message='设备不存在', code=404)

    return api_response(data=dict(row))


@app.route('/api/v1/devices', methods=['POST'])
@require_api_key('write')
def v1_create_device():
    """创建设备记录

    请求体 (JSON):
        department_id  - 单位ID (必填)
        user_name      - 使用人 (必填)
        user_phone     - 联系电话
        user_position  - 安装位置
        computer_name  - 电脑名称
        ip_address     - IP地址
        mac_address    - MAC地址
        dhcp_enabled   - DHCP (是/否)
        os_info        - 操作系统
        cpu_info       - CPU
        ram_info       - 内存
        disk_info      - 硬盘
        motherboard_info - 主板
        gpu_info       - 显卡
        network_adapter - 网卡
        subnet_mask    - 子网掩码
        gateway        - 默认网关
        dns_servers    - DNS服务器
        force          - 强制提交(忽略IP/MAC重复)
    """
    data = request.json
    required = ['department_id', 'user_name']
    for field in required:
        if not data.get(field):
            return api_response(message=f'缺少必填字段: {field}', code=400)

    ip_address = data.get('ip_address', '')
    mac_address = data.get('mac_address', '')
    force = data.get('force', False)
    key_name = g.api_key_info.get('name', 'API')

    conn = get_db()

    # 检查 IP/MAC 重复
    duplicates = []
    if ip_address and ip_address not in ('未知', '0.0.0.0'):
        ip_dup = conn.execute('''
            SELECT d.id, d.computer_name, d.user_name, d.ip_address,
                   dept.name as department_name
            FROM devices d
            LEFT JOIN departments dept ON d.department_id = dept.id
            WHERE d.ip_address = ? AND d.ip_address != '未知'
        ''', (ip_address,)).fetchall()
        for r in ip_dup:
            duplicates.append({
                'type': 'IP地址重复', 'field': 'ip_address', 'value': ip_address,
                'device_id': r['id'], 'computer_name': r['computer_name'],
                'user_name': r['user_name'], 'department_name': r['department_name'],
            })

    if mac_address and mac_address not in ('未知', '00:00:00:00:00:00'):
        mac_dup = conn.execute('''
            SELECT d.id, d.computer_name, d.user_name, d.mac_address,
                   dept.name as department_name
            FROM devices d
            LEFT JOIN departments dept ON d.department_id = dept.id
            WHERE d.mac_address = ? AND d.mac_address != '未知'
        ''', (mac_address,)).fetchall()
        for r in mac_dup:
            duplicates.append({
                'type': 'MAC地址重复', 'field': 'mac_address', 'value': mac_address,
                'device_id': r['id'], 'computer_name': r['computer_name'],
                'user_name': r['user_name'], 'department_name': r['department_name'],
            })

    if duplicates and not force:
        conn.close()
        add_log('DUPLICATE_WARNING',
                f'API提交检测到IP/MAC重复（未确认）',
                f'IP:{ip_address}, MAC:{mac_address}, API Key:{key_name}',
                operator=f'API:{key_name}')
        return api_response(data={'duplicates': duplicates},
                            message='检测到IP或MAC地址与已有设备重复，设置 force=true 强制提交',
                            code=409)

    if duplicates and force:
        add_log('DUPLICATE_CONFIRMED',
                f'API提交（确认覆盖重复）',
                f'IP:{ip_address}, MAC:{mac_address}, API Key:{key_name}',
                operator=f'API:{key_name}')

    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO devices (
            department_id, user_name, user_phone, user_position,
            computer_name, ip_address, mac_address, dhcp_enabled,
            os_info, cpu_info, ram_info, disk_info,
            motherboard_info, gpu_info, network_adapter,
            subnet_mask, gateway, dns_servers
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data.get('department_id'), data.get('user_name'), data.get('user_phone'),
        data.get('user_position'), data.get('computer_name'), data.get('ip_address'),
        data.get('mac_address'), data.get('dhcp_enabled'), data.get('os_info'),
        data.get('cpu_info'), data.get('ram_info'), data.get('disk_info'),
        data.get('motherboard_info'), data.get('gpu_info'), data.get('network_adapter'),
        data.get('subnet_mask'), data.get('gateway'), data.get('dns_servers'),
    ))
    conn.commit()
    device_id = cursor.lastrowid
    conn.close()

    add_log('DEVICE_SUBMIT',
            f'API 设备信息提交成功',
            f'设备ID:{device_id}, IP:{ip_address}, MAC:{mac_address}, API Key:{key_name}',
            operator=f'API:{key_name}')

    return api_response(data={'id': device_id}, message='创建成功', code=201)


@app.route('/api/v1/devices/<int:device_id>', methods=['PUT'])
@require_api_key('write')
def v1_update_device(device_id):
    """更新设备记录"""
    data = request.json
    conn = get_db()

    editable_fields = [
        'department_id', 'user_name', 'user_phone', 'user_position',
        'computer_name', 'ip_address', 'mac_address', 'dhcp_enabled',
        'os_info', 'cpu_info', 'ram_info', 'disk_info',
        'motherboard_info', 'gpu_info', 'network_adapter',
        'subnet_mask', 'gateway', 'dns_servers'
    ]

    updates = []
    values = []
    for field in editable_fields:
        if field in data:
            updates.append(f'{field} = ?')
            values.append(data[field])

    if not updates:
        return api_response(message='没有需要更新的字段', code=400)

    values.append(device_id)
    sql = f"UPDATE devices SET {', '.join(updates)} WHERE id = ?"
    result = conn.execute(sql, values)
    conn.commit()

    if result.rowcount == 0:
        conn.close()
        return api_response(message='设备不存在', code=404)

    key_name = g.api_key_info.get('name', 'API')
    add_log('DEVICE_EDIT',
            f'API 设备记录编辑 (ID:{device_id})',
            f'修改字段: {", ".join(f for f in editable_fields if f in data)}, API Key:{key_name}',
            operator=f'API:{key_name}',
            ip_address=request.remote_addr)

    conn.close()
    return api_response(message='更新成功')


@app.route('/api/v1/devices/<int:device_id>', methods=['DELETE'])
@require_api_key('write')
def v1_delete_device(device_id):
    """删除设备记录"""
    conn = get_db()
    row = conn.execute(
        'SELECT d.*, dept.name as department_name FROM devices d LEFT JOIN departments dept ON d.department_id = dept.id WHERE d.id = ?',
        (device_id,)
    ).fetchone()

    if not row:
        conn.close()
        return api_response(message='设备不存在', code=404)

    device_info = dict(row)
    conn.execute('DELETE FROM devices WHERE id = ?', (device_id,))
    conn.commit()
    conn.close()

    key_name = g.api_key_info.get('name', 'API')
    add_log('DEVICE_DELETE',
            f'API 设备记录删除 (ID:{device_id})',
            f'电脑:{device_info.get("computer_name","?")}, IP:{device_info.get("ip_address","?")}',
            operator=f'API:{key_name}',
            ip_address=request.remote_addr)

    return api_response(message='删除成功')


@app.route('/api/v1/devices/check-duplicates', methods=['GET'])
@require_api_key('read')
def v1_check_duplicates():
    """检查 IP/MAC 重复"""
    # 复用原有逻辑
    conn = get_db()
    ip_dups = conn.execute('''
        SELECT d1.id, d1.computer_name, d1.user_name, d1.ip_address, d1.mac_address,
               dept.name as department_name, p.name as project_name
        FROM devices d1
        LEFT JOIN departments dept ON d1.department_id = dept.id
        LEFT JOIN projects p ON dept.project_id = p.id
        INNER JOIN devices d2 ON d1.ip_address = d2.ip_address AND d1.id != d2.id
        WHERE d1.ip_address IS NOT NULL AND d1.ip_address != '' AND d1.ip_address != '未知'
        GROUP BY d1.ip_address, d1.id ORDER BY d1.ip_address
    ''').fetchall()
    mac_dups = conn.execute('''
        SELECT d1.id, d1.computer_name, d1.user_name, d1.ip_address, d1.mac_address,
               dept.name as department_name, p.name as project_name
        FROM devices d1
        LEFT JOIN departments dept ON d1.department_id = dept.id
        LEFT JOIN projects p ON dept.project_id = p.id
        INNER JOIN devices d2 ON d1.mac_address = d2.mac_address AND d1.id != d2.id
        WHERE d1.mac_address IS NOT NULL AND d1.mac_address != '' AND d1.mac_address != '未知'
        GROUP BY d1.mac_address, d1.id ORDER BY d1.mac_address
    ''').fetchall()
    conn.close()

    ip_groups = {}
    for r in ip_dups:
        ip_groups.setdefault(r['ip_address'], []).append(dict(r))
    mac_groups = {}
    for r in mac_dups:
        mac_groups.setdefault(r['mac_address'], []).append(dict(r))

    return api_response(data={
        'ip_duplicates': ip_groups,
        'mac_duplicates': mac_groups,
        'ip_duplicate_count': len(ip_groups),
        'mac_duplicate_count': len(mac_groups),
        'total_duplicate_devices': len(set(
            [r['id'] for r in ip_dups] + [r['id'] for r in mac_dups]
        ))
    })


@app.route('/api/v1/projects', methods=['GET'])
@require_api_key('read')
def v1_get_projects():
    """获取项目列表"""
    conn = get_db()
    rows = conn.execute('''
        SELECT p.*,
            (SELECT COUNT(*) FROM users WHERE project_id = p.id) as user_count,
            (SELECT COUNT(*) FROM departments WHERE project_id = p.id) as dept_count,
            (SELECT COUNT(*) FROM devices d JOIN departments dept ON d.department_id = dept.id WHERE dept.project_id = p.id) as device_count
        FROM projects p ORDER BY p.id
    ''').fetchall()
    conn.close()
    return api_response(data=[dict(r) for r in rows])


@app.route('/api/v1/departments', methods=['GET'])
@require_api_key('read')
def v1_get_departments():
    """获取单位列表"""
    project_id = request.args.get('project_id')
    conn = get_db()
    query = 'SELECT d.*, p.name as project_name FROM departments d LEFT JOIN projects p ON d.project_id = p.id WHERE 1=1'
    params = []
    if project_id:
        query += ' AND d.project_id = ?'
        params.append(project_id)
    query += ' ORDER BY d.id'
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return api_response(data=[dict(r) for r in rows])


@app.route('/api/v1/stats', methods=['GET'])
@require_api_key('read')
def v1_get_stats():
    """获取统计信息"""
    conn = get_db()
    project_count = conn.execute('SELECT COUNT(*) FROM projects').fetchone()[0]
    user_count = conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]
    dept_count = conn.execute('SELECT COUNT(*) FROM departments').fetchone()[0]
    device_count = conn.execute('SELECT COUNT(*) FROM devices').fetchone()[0]
    today_count = conn.execute(
        "SELECT COUNT(*) FROM devices WHERE DATE(collected_at) = DATE('now')"
    ).fetchone()[0]
    project_stats = conn.execute('''
        SELECT p.name, COALESCE(COUNT(d.id), 0) as count
        FROM projects p
        LEFT JOIN departments dept ON dept.project_id = p.id
        LEFT JOIN devices d ON d.department_id = dept.id
        GROUP BY p.id ORDER BY count DESC
    ''').fetchall()
    conn.close()
    return api_response(data={
        'project_count': project_count,
        'user_count': user_count,
        'department_count': dept_count,
        'device_count': device_count,
        'today_count': today_count,
        'project_stats': [dict(r) for r in project_stats]
    })


@app.route('/api/v1/export/<format>', methods=['GET'])
@require_api_key('read')
def v1_export_devices(format):
    """导出设备列表 (xlsx/csv)

    查询参数:
        project_id    - 按项目筛选
        department_id - 按单位筛选
        keyword       - 关键词搜索
    """
    if format not in ('xlsx', 'csv'):
        return api_response(message='不支持的格式，请使用 xlsx 或 csv', code=400)

    # 复用原导出逻辑
    project_id = request.args.get('project_id')
    dept_id = request.args.get('department_id')
    keyword = request.args.get('keyword', '').strip()

    conn = get_db()
    query = '''
        SELECT d.*, dept.name as department_name, dept.code as department_code,
               dept.project_id, p.name as project_name
        FROM devices d
        LEFT JOIN departments dept ON d.department_id = dept.id
        LEFT JOIN projects p ON dept.project_id = p.id
        WHERE 1=1
    '''
    params = []
    if project_id:
        query += ' AND dept.project_id = ?'
        params.append(project_id)
    if dept_id:
        query += ' AND d.department_id = ?'
        params.append(dept_id)
    if keyword:
        query += ' AND (d.user_name LIKE ? OR d.computer_name LIKE ? OR d.ip_address LIKE ?)'
        kw = f'%{keyword}%'
        params.extend([kw, kw, kw])
    query += ' ORDER BY d.collected_at DESC'
    rows = conn.execute(query, params).fetchall()
    conn.close()

    columns = [
        ('ID', 'id'), ('所属项目', 'project_name'), ('所属单位', 'department_name'),
        ('单位编码', 'department_code'), ('使用人', 'user_name'), ('联系电话', 'user_phone'),
        ('安装位置', 'user_position'), ('电脑名称', 'computer_name'), ('IP 地址', 'ip_address'),
        ('MAC 地址', 'mac_address'), ('自动获取IP(DHCP)', 'dhcp_enabled'),
        ('子网掩码', 'subnet_mask'), ('默认网关', 'gateway'), ('DNS 服务器', 'dns_servers'),
        ('网卡', 'network_adapter'), ('操作系统', 'os_info'), ('CPU', 'cpu_info'),
        ('内存', 'ram_info'), ('硬盘', 'disk_info'), ('主板', 'motherboard_info'),
        ('显卡', 'gpu_info'), ('采集时间', 'collected_at'),
    ]

    if format == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([col[0] for col in columns])
        for row in rows:
            r = dict(row)
            writer.writerow([r.get(col[1], '') or '' for col in columns])
        mem = io.BytesIO()
        mem.write(output.getvalue().encode('utf-8-sig'))
        mem.seek(0)
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(mem, mimetype='text/csv', as_attachment=True,
                         download_name=f'设备列表_{timestamp}.csv')
    else:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '设备列表'

        header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='1A73E8', end_color='1A73E8', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'),
        )
        for col_idx, (header, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        data_font = Font(name='微软雅黑', size=10)
        data_align = Alignment(vertical='center', wrap_text=True)
        for row_idx, row in enumerate(rows, 2):
            r = dict(row)
            for col_idx, (_, key) in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=r.get(key, '') or '')
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

        mem = io.BytesIO()
        wb.save(mem)
        mem.seek(0)
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(mem,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'设备列表_{timestamp}.xlsx')


# ==================== 页面路由 ====================

@app.route('/')
def index():
    """管理端主页 - 需要登录"""
    if 'user_id' not in session:
        return redirect(url_for('login_page'))
    return send_from_directory('templates', 'index.html')


@app.route('/login')
def login_page():
    """登录页面"""
    if 'user_id' in session:
        return redirect(url_for('index'))
    return send_from_directory('templates', 'login.html')


@app.route('/api-docs')
def api_docs():
    """API 文档页面"""
    return send_from_directory('templates', 'api_docs.html')


# ==================== 生成客户端 ====================

def _build_client_package(user, server_url, plain_password, pack_mode='zip'):
    """
    构建客户端包（内部函数）
    pack_mode: 'zip' = Python源码+BAT, 'exe' = PyInstaller打包的EXE
    返回 (zip_path, zip_filename)
    """
    # 客户端源文件目录
    client_src = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'client')
    if not os.path.exists(client_src):
        raise Exception('客户端源文件不存在，请检查 client/ 目录')

    # 加密密码
    encrypted_pwd = aes_encrypt(plain_password)

    # 创建临时目录
    tmp_dir = tempfile.mkdtemp(prefix='dc_client_')

    try:
        # 复制客户端文件（排除 CONFIG.INI 模板）
        for fname in os.listdir(client_src):
            if fname.lower() == 'config.ini':
                continue  # 不复制模板，后面自己生成
            src = os.path.join(client_src, fname)
            dst = os.path.join(tmp_dir, fname)
            if os.path.isfile(src):
                shutil.copy2(src, dst)

        # 生成 CONFIG.INI（密码为密文）
        config_content = f"""# ============================================
# 设备信息采集器 - 客户端配置文件
# ============================================
# 由服务端自动生成，包含服务器地址和加密登录凭据
# 生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
# 关联用户: {user['display_name'] or user['username']}
# 关联项目: {user['project_name'] or '未关联'}

[Server]
# 服务器地址
ServerUrl = {server_url}

[Account]
# 登录账号
Username = {user['username']}

# 登录密码（AES加密密文，请勿修改）
Password = ENC:{encrypted_pwd}
"""
        config_path = os.path.join(tmp_dir, 'CONFIG.INI')
        with open(config_path, 'w', encoding='utf-8') as f:
            f.write(config_content)

        if pack_mode == 'exe':
            # 优先检测预编译 EXE（Docker 容器中的 Windows 交叉编译产物）
            prebuilt_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'client', 'prebuilt')
            prebuilt_exe = os.path.join(prebuilt_dir, 'DeviceCollector.exe')

            if os.path.exists(prebuilt_exe):
                # ===== 预编译 EXE 模式（无需 PyInstaller，直接复制 + 替换 CONFIG） =====
                app.logger.info(f"[客户端生成] 使用预编译EXE: {prebuilt_exe}")

                display_name = user['display_name'] or user['username']
                safe_name = user['username'].replace(' ', '_').replace('.', '_')
                display_exe_name = f"设备采集器_{display_name}"

                # 创建打包目录
                pkg_dir = os.path.join(tmp_dir, 'package')
                os.makedirs(pkg_dir, exist_ok=True)

                # 复制预编译 EXE 并重命名
                dst_exe = os.path.join(pkg_dir, f'{display_exe_name}.exe')
                shutil.copy2(prebuilt_exe, dst_exe)

                # 复制 CONFIG.INI
                shutil.copy2(config_path, os.path.join(pkg_dir, 'CONFIG.INI'))

                # 写使用说明
                readme = f"""设备采集器 - 使用说明
============================================

文件说明:
  {display_exe_name}.exe  - 客户端程序，双击运行
  CONFIG.INI              - 配置文件（含服务器地址和加密登录凭据）

使用方法:
  1. 解压后保持所有文件在同一文件夹中
  2. 双击 {display_exe_name}.exe 运行
  3. 程序会自动读取 CONFIG.INI 中的配置并登录

注意事项:
  - CONFIG.INI 中的密码为 AES 加密密文，请勿手动修改
  - 如需更改服务器地址或账号，请联系管理员重新生成
  - 整个文件夹可以复制到其他电脑使用，无需安装 Python

生成信息:
  生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
  关联用户: {display_name}
  关联项目: {user['project_name'] or '未关联'}
  服务器:   {server_url}
"""
                with open(os.path.join(pkg_dir, '使用说明.txt'), 'w', encoding='utf-8') as f:
                    f.write(readme)

                # 打包为 ZIP
                zip_filename = f"设备采集客户端_{display_name}_EXE版.zip"
                zip_path = os.path.join(tempfile.gettempdir(),
                                        f'dc_client_{user["id"]}_{int(time.time())}.zip')
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for root, dirs, files in os.walk(pkg_dir):
                        for fname in files:
                            fpath = os.path.join(root, fname)
                            arcname = os.path.relpath(fpath, pkg_dir)
                            zf.write(fpath, arcname)

                # 清理
                shutil.rmtree(tmp_dir, ignore_errors=True)
                app.logger.info(f"[客户端生成] 预编译EXE打包成功: {zip_filename}")
                return zip_path, zip_filename

            elif sys.platform.startswith('linux'):
                # Linux 环境且无预编译 EXE，回退到 zip 模式
                app.logger.warning("[客户端生成] Linux 环境无预编译EXE，自动切换为 zip 模式")
                pack_mode = 'zip'

        if pack_mode == 'exe':
            # ===== PyInstaller 打包模式（onedir — 不触发杀毒误报） =====
            python_exe = shutil.which('python') or 'python'
            # 尝试用 venv 中的 python
            venv_python = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))),
                                       '.workbuddy', 'binaries', 'python', 'envs', 'default', 'Scripts', 'python.exe')
            if not os.path.exists(venv_python):
                venv_python2 = os.path.join(os.path.expanduser('~'),
                                            '.workbuddy', 'binaries', 'python', 'envs', 'default', 'Scripts', 'python.exe')
                if os.path.exists(venv_python2):
                    venv_python = venv_python2
            if os.path.exists(venv_python):
                python_exe = venv_python

            # 检查 PyInstaller 是否可用
            pi_check = sp.run([python_exe, '-m', 'PyInstaller', '--version'],
                              capture_output=True, text=True, timeout=10)
            if pi_check.returncode != 0:
                shutil.rmtree(tmp_dir, ignore_errors=True)
                raise Exception('PyInstaller 未安装，无法生成 EXE。请在服务器端安装: pip install pyinstaller')

            client_py = os.path.join(tmp_dir, 'client.py')

            # PyInstaller --name 用英文
            safe_name = user['username'].replace(' ', '_').replace('.', '_')
            pyinstaller_name = f"DeviceCollector_{safe_name}"
            display_exe_name = f"设备采集器_{user['display_name'] or user['username']}"

            dist_dir = os.path.join(tmp_dir, 'dist')
            build_dir = os.path.join(tmp_dir, 'build')

            # 定位客户端资源文件（版本信息 + 应用清单）
            client_res_dir = os.path.join(os.path.dirname(__file__), '..', 'client')
            version_file = os.path.join(client_res_dir, 'version_info.py')
            manifest_file = os.path.join(client_res_dir, 'app.manifest')

            # PyInstaller 命令 — onedir 模式（避免杀毒误报）
            cmd = [
                python_exe, '-m', 'PyInstaller',
                '--onedir',                     # 目录模式（不触发杀毒误报）
                '--windowed',                   # 无控制台窗口（GUI程序不需要）
                '--noconfirm',                  # 不询问覆盖
                '--clean',                      # 清理临时文件
                '--name', pyinstaller_name,      # EXE 文件名（必须英文）
                '--distpath', dist_dir,
                '--workpath', build_dir,
                '--specpath', tmp_dir,
                '--hidden-import', 'Crypto.Cipher.AES',
                '--hidden-import', 'Crypto.Util.Padding',
                '--hidden-import', 'Crypto.Cipher',
                '--hidden-import', 'Crypto.Util',
            ]

            # 添加版本信息（减少杀毒误报）
            if os.path.exists(version_file):
                cmd.extend(['--version-file', version_file])
                app.logger.info(f"[EXE打包] 包含版本信息: {version_file}")

            # 添加应用清单（UAC + 兼容性）
            if os.path.exists(manifest_file):
                cmd.extend(['--manifest', manifest_file])
                app.logger.info(f"[EXE打包] 包含应用清单: {manifest_file}")

            cmd.append(client_py)

            app.logger.info(f"[EXE打包] 开始打包: {' '.join(cmd)}")
            result = sp.run(cmd, capture_output=True, text=True, timeout=300, cwd=tmp_dir)

            if result.returncode != 0:
                err_detail = result.stderr or result.stdout or '未知错误'
                err_lines = [l for l in err_detail.split('\n')
                             if 'error' in l.lower() or 'failed' in l.lower() or 'ERROR' in l]
                err_summary = err_lines[-1] if err_lines else err_detail[:500]
                app.logger.error(f"[EXE打包] 失败: {err_summary}")
                shutil.rmtree(tmp_dir, ignore_errors=True)
                raise Exception(f'EXE 打包失败: {err_summary}')

            # onedir 模式：EXE 在 dist/pyinstaller_name/ 目录下
            onedir_path = os.path.join(dist_dir, pyinstaller_name)
            exe_path = os.path.join(onedir_path, f'{pyinstaller_name}.exe')
            if not os.path.exists(exe_path):
                # 列出 dist 内容帮助排查
                dist_files = []
                if os.path.exists(dist_dir):
                    for root, dirs, files in os.walk(dist_dir):
                        for f in files:
                            dist_files.append(os.path.relpath(os.path.join(root, f), dist_dir))
                app.logger.error(f"[EXE打包] 未找到EXE: {exe_path}, dist内容: {dist_files}")
                shutil.rmtree(tmp_dir, ignore_errors=True)
                raise Exception(f'EXE 文件未生成 (dist内容: {dist_files[:10]})')

            # 将整个 onedir 目录复制为最终打包目录
            # EXE 重命名为中文显示名
            pkg_dir = os.path.join(tmp_dir, 'package')
            shutil.copytree(onedir_path, pkg_dir)

            # 重命名 EXE 为中文显示名
            old_exe = os.path.join(pkg_dir, f'{pyinstaller_name}.exe')
            new_exe = os.path.join(pkg_dir, f'{display_exe_name}.exe')
            if os.path.exists(old_exe):
                os.rename(old_exe, new_exe)

            # 添加 CONFIG.INI
            shutil.copy2(config_path, os.path.join(pkg_dir, 'CONFIG.INI'))

            # 添加使用说明
            readme = f"""设备采集器 - 使用说明
============================================

文件说明:
  {display_exe_name}.exe  - 客户端程序，双击运行
  CONFIG.INI              - 配置文件（含服务器地址和加密登录凭据）
  其他文件                - 程序运行依赖，请勿删除

使用方法:
  1. 解压后保持所有文件在同一文件夹中
  2. 双击 {display_exe_name}.exe 运行
  3. 程序会自动读取 CONFIG.INI 中的配置并登录

注意事项:
  - CONFIG.INI 中的密码为 AES 加密密文，请勿手动修改
  - 如需更改服务器地址或账号，请联系管理员重新生成
  - 请勿删除 CONFIG.INI 和其他依赖文件，否则无法运行
  - 整个文件夹可以复制到其他电脑使用

生成信息:
  生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
  关联用户: {user['display_name'] or user['username']}
  关联项目: {user['project_name'] or '未关联'}
  服务器:   {server_url}
"""
            with open(os.path.join(pkg_dir, '使用说明.txt'), 'w', encoding='utf-8') as f:
                f.write(readme)

            # 打包为 ZIP
            zip_filename = f"设备采集客户端_{user['display_name'] or user['username']}_EXE版.zip"
            zip_path = os.path.join(tempfile.gettempdir(),
                                    f'dc_client_{user["id"]}_{int(time.time())}.zip')
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for root, dirs, files in os.walk(pkg_dir):
                    for fname in files:
                        fpath = os.path.join(root, fname)
                        # ZIP 内使用相对路径保持目录结构
                        arcname = os.path.relpath(fpath, pkg_dir)
                        zf.write(fpath, arcname)

            # 清理
            shutil.rmtree(tmp_dir, ignore_errors=True)
            app.logger.info(f"[EXE打包] 成功: {zip_filename}")
            return zip_path, zip_filename

        if pack_mode == 'zip':
            # ===== ZIP 模式（Python 源码 + BAT 启动脚本） =====

            # 创建启动脚本
            bat_content = f"""@echo off
chcp 65001 >nul
title 设备信息采集器 - {user['display_name'] or user['username']}
echo ============================================
echo   设备信息采集器 - 客户端
echo   用户: {user['display_name'] or user['username']}
echo   项目: {user['project_name'] or '未关联'}
echo ============================================
echo.
cd /d "%~dp0"
python client.py
pause
"""
            bat_path = os.path.join(tmp_dir, '启动客户端.bat')
            with open(bat_path, 'w', encoding='utf-8') as f:
                f.write(bat_content)

            # 打包为 ZIP
            zip_filename = f"设备采集客户端_{user['display_name'] or user['username']}.zip"
            zip_path = os.path.join(tempfile.gettempdir(),
                                    f'dc_client_{user["id"]}_{int(time.time())}.zip')
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for fname in os.listdir(tmp_dir):
                    fpath = os.path.join(tmp_dir, fname)
                    if os.path.isfile(fpath):
                        zf.write(fpath, fname)

            # 清理
            shutil.rmtree(tmp_dir, ignore_errors=True)
            return zip_path, zip_filename

    except Exception:
        if os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir, ignore_errors=True)
        raise


@app.route('/api/generate-client', methods=['POST'])
def generate_client():
    """生成客户端程序（含加密密码的 CONFIG.INI）"""
    # 需要管理员权限
    if not session.get('user_id') or session.get('role') != 'admin':
        return jsonify({'error': '需要管理员权限'}), 403

    data = request.json or {}
    user_id = data.get('user_id')
    server_url = data.get('server_url', '').strip()
    plain_password = data.get('password', '').strip()
    pack_mode = data.get('pack_mode', 'exe')  # 'exe' 或 'zip'

    if not user_id:
        return jsonify({'error': '请选择用户'}), 400
    if not plain_password:
        return jsonify({'error': '请输入用户密码'}), 400

    # 获取用户信息并验证密码
    conn = get_db()
    user = conn.execute(
        'SELECT u.*, p.name as project_name, p.code as project_code '
        'FROM users u LEFT JOIN projects p ON u.project_id = p.id '
        'WHERE u.id = ?', (user_id,)
    ).fetchone()
    conn.close()

    if not user:
        return jsonify({'error': '用户不存在'}), 404

    # 验证密码
    if user['password_hash'] != hash_password(plain_password):
        return jsonify({'error': '密码验证失败，请输入该用户的正确密码'}), 401

    # 如果未提供服务器地址，使用请求来源
    if not server_url:
        server_url = request.host_url.rstrip('/')

    try:
        zip_path, zip_filename = _build_client_package(user, server_url, plain_password, pack_mode)

        # 记录日志
        add_log('CLIENT_GENERATE',
                f'生成客户端: {user["display_name"] or user["username"]} ({pack_mode})',
                f'用户ID:{user_id}, 项目:{user["project_name"] or "无"}, 服务器:{server_url}, 打包:{pack_mode}',
                operator=session.get('username'),
                ip_address=request.remote_addr)

        return send_file(
            zip_path,
            mimetype='application/zip',
            as_attachment=True,
            download_name=zip_filename
        )

    except Exception as e:
        return jsonify({'error': f'生成客户端失败: {str(e)}'}), 500


@app.route('/api/generate-client-with-password', methods=['POST'])
def generate_client_with_password():
    """生成客户端程序（兼容旧接口，默认 exe 模式）"""
    # 需要管理员权限
    if not session.get('user_id') or session.get('role') != 'admin':
        return jsonify({'error': '需要管理员权限'}), 403

    data = request.json or {}
    user_id = data.get('user_id')
    server_url = data.get('server_url', '').strip()
    plain_password = data.get('password', '').strip()

    if not user_id:
        return jsonify({'error': '请选择用户'}), 400
    if not plain_password:
        return jsonify({'error': '请输入用户密码'}), 400

    # 获取用户信息并验证密码
    conn = get_db()
    user = conn.execute(
        'SELECT u.*, p.name as project_name, p.code as project_code '
        'FROM users u LEFT JOIN projects p ON u.project_id = p.id '
        'WHERE u.id = ?', (user_id,)
    ).fetchone()
    conn.close()

    if not user:
        return jsonify({'error': '用户不存在'}), 404

    # 验证密码
    if user['password_hash'] != hash_password(plain_password):
        return jsonify({'error': '密码验证失败，请输入该用户的正确密码'}), 401

    # 如果未提供服务器地址，使用请求来源
    if not server_url:
        server_url = request.host_url.rstrip('/')

    try:
        zip_path, zip_filename = _build_client_package(user, server_url, plain_password, 'exe')

        # 记录日志
        add_log('CLIENT_GENERATE',
                f'生成客户端: {user["display_name"] or user["username"]} (exe)',
                f'用户ID:{user_id}, 项目:{user["project_name"] or "无"}, 服务器:{server_url}',
                operator=session.get('username'),
                ip_address=request.remote_addr)

        return send_file(
            zip_path,
            mimetype='application/zip',
            as_attachment=True,
            download_name=zip_filename
        )

    except Exception as e:
        return jsonify({'error': f'生成客户端失败: {str(e)}'}), 500


if __name__ == '__main__':
    init_db()
    print('=' * 50)
    print('  设备信息采集器 - 服务端已启动')
    print(f'  数据库: {DB_PATH}')
    print('  管理界面: http://0.0.0.0:5000')
    print('  API 文档: http://0.0.0.0:5000/api-docs')
    print('  默认管理员: admin / 123456')
    print('=' * 50)
    app.run(host='0.0.0.0', port=5000, debug=True)
